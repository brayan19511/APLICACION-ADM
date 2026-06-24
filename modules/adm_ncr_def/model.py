from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from core.business_config import DefRules
from core.dataframe_schema import (
    clean_text,
    normalize_columns,
    numeric_series,
    require_columns,
)
from core.exceptions import ExportError, ValidationError


DEF_ALIASES = {
    "N° NOTA CRÉDITO": (
        "NRO NOTA CREDITO",
        "NUMERO NOTA CREDITO",
        "NOTA CREDITO",
    ),
    "FECHA NCR": ("FECHA NOTA CREDITO",),
    "TIPO DOC.": ("TIPO DOC", "TIPO DOCUMENTO"),
    "N° DOCUMENTO": ("NRO DOCUMENTO", "NUMERO DOCUMENTO", "DOCUMENTO","Doc. Cliente"),
    "DNI/RUC": ("DNI RUC", "DOCUMENTO IDENTIDAD"),
    "NOMBRE CLIENTE/RAZÓN SOCIAL": (
        "NOMBRE CLIENTE RAZON SOCIAL",
        "CLIENTE",
        "RAZON SOCIAL",
        "Nombre / Razon Social"
    ),
    "IMPORTE": ("MONTO", "MONTO NCR"),
    "FECHA REGISTRO DATOS": ("FECHA REGISTRO"
                            #  TODO:VALIDAR
                            ),
    "CUENTA TITULAR": ("TITULAR CUENTA",),
    "N° CUENTA": ("NRO CUENTA", "NUMERO CUENTA", "CUENTA BANCARIA","N° de Cuenta"),
    "CCI": ("CODIGO CUENTA INTERBANCARIO", "CUENTA INTERBANCARIA"),
    "BANCO": ("ENTIDAD BANCARIA",),
    "ESTADO": ("STATUS",),
    "CORREO": ("EMAIL", "CORREO ELECTRONICO"),
    "DEF FORMATEADO": ("DEF", "NCR FORMATEADA"),
    "Doc. Verificar": ("DOC VERIFICAR", "DOCUMENTO VERIFICAR"),
    "Clasificacion Doc": (
        "CLASIFICACION DOCUMENTO",
        "TIPO DOCUMENTO BCP",
    ),
    "Clasificacion Banco": ("CLASIFICACION BANCARIA", "TIPO CUENTA BCP"),
    "Cuenta Seleccionada": ("CUENTA SELECCIONADA", "CUENTA ABONO"),
}

DEF_REQUIRED_COLUMNS = (
    "N° NOTA CRÉDITO",
    "FECHA NCR",
    "TIPO DOC.",
    "N° DOCUMENTO",
    "DNI/RUC",
    "NOMBRE CLIENTE/RAZÓN SOCIAL",
    "IMPORTE",
    "FECHA REGISTRO DATOS",
    "CUENTA TITULAR",
    "N° CUENTA",
    "CCI",
    "BANCO",
    "ESTADO",
    "CORREO",
)

BCP_REQUIRED_COLUMNS = (
    "Clasificacion Banco",
    "Cuenta Seleccionada",
    "Clasificacion Doc",
    "DNI/RUC",
    "NOMBRE CLIENTE/RAZÓN SOCIAL",
    "IMPORTE",
    "DEF FORMATEADO",
)

TEXT_COLUMNS = (
    "N° NOTA CRÉDITO",
    "TIPO DOC.",
    "N° DOCUMENTO",
    "DNI/RUC",
    "N° CUENTA",
    "CCI",
    "BANCO",
    "ESTADO",
)


def _read_excel(file_path: str | Path) -> pd.DataFrame:
    path = Path(file_path) if file_path else None
    if path is None:
        raise ValidationError("No se ha seleccionado un archivo.")
    if not path.is_file():
        raise FileNotFoundError(f"El archivo '{path}' no existe.")
    try:
        return normalize_columns(pd.read_excel(path, dtype=object), DEF_ALIASES)
    except (ValidationError, FileNotFoundError):
        raise
    except Exception as exc:
        raise ValidationError(f"No se pudo leer el archivo Excel: {exc}") from exc


def _move_after(dataframe: pd.DataFrame, column: str, after: str) -> None:
    value = dataframe.pop(column)
    position = dataframe.columns.get_loc(after) + 1
    dataframe.insert(position, column, value)


class NcrDefProcesor:
    def __init__(self, path=None, rules: DefRules | None = None):
        self.path = path
        self.df: pd.DataFrame | None = None
        self.rules = rules or DefRules()

    def getFilePath(self):
        return self.path

    def getDf(self):
        return self.df

    def CargarDataFrame(self, file_path):
        self.df = _read_excel(file_path)
        self.path = str(Path(file_path).resolve())

    def validate_columns_DEF(self):
        require_columns(self.df, DEF_REQUIRED_COLUMNS)

    def procesarDEF(self):
        self.validate_columns_DEF()
        assert self.df is not None

        for column in TEXT_COLUMNS:
            self.df[column] = clean_text(self.df[column])
        self.df["IMPORTE"] = numeric_series(self.df["IMPORTE"], "IMPORTE")

        status = self.df["ESTADO"].str.upper()
        self.df = self.df[status.eq(self.rules.accepted_status)].copy()
        self.clasificacion()
        self.ajustarColumnas()

    def clasificacion(self):
        assert self.df is not None
        self.df["DEF FORMATEADO"] = (
            clean_text(self.df["N° NOTA CRÉDITO"]).str.replace("-", "", regex=False)
        )
        self.df["DNI/RUC"] = self.df["DNI/RUC"].apply(
            lambda value: value[2:-1]
            if len(value) == 11 and value.startswith("10")
            else value
        )
        self.df["DNI/RUC"] = self.df.apply(
            lambda row: self.ajusteCEX(row["DNI/RUC"], row["TIPO DOC."]), axis=1
        )
        self.df["N° DOCUMENTO"] = self.df.apply(
            lambda row: self.ajusteCEX(row["N° DOCUMENTO"], row["TIPO DOC."]),
            axis=1,
        )
        self.df["Doc. Verificar"] = np.where(
            self.df["DNI/RUC"].eq(self.df["N° DOCUMENTO"]), "ok", "revisar"
        )
        self.df["Clasificacion Doc"] = self.df["DNI/RUC"].apply(
            self.clasificacionDoc
        )

        bank = clean_text(self.df["BANCO"]).str.upper()
        for alias in self.rules.bcp_aliases:
            bank = bank.replace(alias, self.rules.bcp_name)
        self.df["BANCO"] = bank
        self.df["Clasificacion Banco"] = self.df.apply(
            lambda row: self.clasificacionBanco(
                row["BANCO"], row["N° CUENTA"], row["CCI"]
            ),
            axis=1,
        )
        self.df["Cuenta Seleccionada"] = np.where(
            self.df["BANCO"].eq(self.rules.bcp_name),
            self.df["N° CUENTA"],
            self.df["CCI"],
        )

    def clasificacionDoc(self, value):
        length = len(str(value))
        return {8: "1", 9: "3", 11: "6"}.get(length, "4")

    def clasificacionBanco(self, banco, cuenta, cci):
        bank = str(banco).strip().upper()
        account = str(cuenta).strip()
        interbank = str(cci).strip()
        if bank == self.rules.bcp_name:
            if len(account) == 13:
                return "C"
            if len(account) == 14:
                return "A"
            return f"revisar BCP: {len(account)} dígitos"
        if len(interbank) == 20:
            return "B"
        return f"revisar CCI: {len(interbank)} dígitos"

    def ajusteCEX(self, doc, tipoDoc):
        document = str(doc).strip()
        if str(tipoDoc).strip().startswith("3"):
            if len(document) < 9:
                return document.zfill(9)
            if len(document) > 9:
                return document[-9:]
        return document

    def ajustarColumnas(self):
        assert self.df is not None
        _move_after(self.df, "DEF FORMATEADO", "N° NOTA CRÉDITO")
        _move_after(self.df, "Doc. Verificar", "DNI/RUC")
        _move_after(self.df, "Clasificacion Doc", "Doc. Verificar")
        _move_after(self.df, "Clasificacion Banco", "BANCO")
        _move_after(self.df, "Cuenta Seleccionada", "Clasificacion Banco")

    def analizar_dataframe_def(self):
        if self.df is None:
            raise ValidationError("Aún no se ha procesado un archivo.")
        review = self.df["Doc. Verificar"].str.contains("revisar", na=False) | (
            self.df["Clasificacion Banco"].str.contains("revisar", na=False)
        )
        return (
            len(self.df),
            self.df["IMPORTE"].sum(),
            int(review.sum()),
            self.df.loc[review, "IMPORTE"].sum(),
            int(self.df["IMPORTE"].gt(self.rules.unusual_amount).sum()),
        )

    def ExportarDataFrame_def(self, file_path="export.xlsx"):
        if self.df is None:
            raise ValidationError("Aún no se ha procesado un archivo.")
        if not file_path:
            raise ValidationError("No se ha seleccionado una ruta de salida.")
        try:
            self.df.to_excel(file_path, index=False)
            workbook = load_workbook(file_path)
            sheet = workbook.active
            fill = PatternFill(
                start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
            )
            for name in ("Doc. Verificar", "Clasificacion Banco"):
                sheet.cell(
                    row=1, column=self.df.columns.get_loc(name) + 1
                ).fill = fill
            for number, cells in enumerate(sheet.iter_cols(), start=1):
                maximum = max(
                    (len(str(cell.value)) for cell in cells if cell.value is not None),
                    default=0,
                )
                sheet.column_dimensions[get_column_letter(number)].width = min(
                    maximum + 2, 60
                )
            workbook.save(file_path)
            workbook.close()
            return f"✅ Archivo guardado en: {file_path}"
        except Exception as exc:
            raise ExportError(f"No se pudo exportar el archivo DEF: {exc}") from exc


class NcrBCPProcesor:
    def __init__(self, path=None, rules: DefRules | None = None):
        self.path = path
        self.df: pd.DataFrame | None = None
        self.df_bcp: pd.DataFrame | None = None
        self.rules = rules or DefRules()

    def getFilePath(self):
        return self.path

    def getDf(self):
        return self.df_bcp

    def CargarDataFrame(self, file_path):
        self.df = _read_excel(file_path)
        self.path = str(Path(file_path).resolve())

    def validate_columns_BCP(self):
        require_columns(self.df, BCP_REQUIRED_COLUMNS)

    def procesarBCP(self):
        self.validate_columns_BCP()
        assert self.df is not None
        for column in (
            "Clasificacion Banco",
            "Cuenta Seleccionada",
            "Clasificacion Doc",
            "DNI/RUC",
            "DEF FORMATEADO",
        ):
            self.df[column] = clean_text(self.df[column])
        self.df["IMPORTE"] = numeric_series(self.df["IMPORTE"], "IMPORTE")
        if "ESTADO" in self.df.columns:
            status = clean_text(self.df["ESTADO"]).str.upper()
            self.df = self.df[status.eq(self.rules.accepted_status)].copy()
        if "Doc. Verificar" not in self.df.columns:
            self.df["Doc. Verificar"] = "ok"
        self.Plantilla()

    def Plantilla(self) -> pd.DataFrame:
        assert self.df is not None
        rows = []
        for item in self.df.to_dict("records"):
            rows.append(
                {
                    "Tipo de Registro": "A",
                    "Tipo de Cuenta de Abono": item["Clasificacion Banco"],
                    "Cuenta de Abono": item["Cuenta Seleccionada"],
                    "Tipo de Documento de Identidad": item["Clasificacion Doc"],
                    "Número de Documento de Identidad": item["DNI/RUC"],
                    "Correlativo de Documento de Identidad": "",
                    "Nombre del proveedor": item["NOMBRE CLIENTE/RAZÓN SOCIAL"],
                    "Tipo de Moneda de Abono": "S",
                    "Monto del Abono": item["IMPORTE"],
                    "Validación IDC del proveedor vs Cuenta": "S",
                    "Cantidad Documentos relacionados al Abono": "0001",
                    "Tipo de Documento a pagar": "",
                    "Nro. del Documento": "",
                    "Moneda Documento": "",
                    "Monto del Documento": "",
                }
            )
            rows.append(
                {
                    "Tipo de Registro": "D",
                    "Tipo de Cuenta de Abono": "",
                    "Cuenta de Abono": "",
                    "Tipo de Documento de Identidad": "",
                    "Número de Documento de Identidad": "",
                    "Correlativo de Documento de Identidad": "",
                    "Nombre del proveedor": "",
                    "Tipo de Moneda de Abono": "",
                    "Monto del Abono": "",
                    "Validación IDC del proveedor vs Cuenta": "",
                    "Cantidad Documentos relacionados al Abono": "",
                    "Tipo de Documento a pagar": "C",
                    "Nro. del Documento": item["DEF FORMATEADO"],
                    "Moneda Documento": "S",
                    "Monto del Documento": item["IMPORTE"],
                }
            )
        self.df_bcp = pd.DataFrame(rows)
        return self.df_bcp

    def analizar_dataframe_BCP(self):
        if self.df is None:
            raise ValidationError("Aún no se ha procesado un archivo.")
        review = clean_text(self.df["Doc. Verificar"]).str.contains(
            "revisar", case=False, na=False
        ) | clean_text(self.df["Clasificacion Banco"]).str.contains(
            "revisar", case=False, na=False
        )
        return (
            len(self.df),
            self.df["IMPORTE"].sum(),
            int(review.sum()),
            self.df.loc[review, "IMPORTE"].sum(),
            int(self.df["IMPORTE"].gt(self.rules.unusual_amount).sum()),
        )

    def ExportarDataFrame_bcp(self, file_path="export.xlsx"):
        if not file_path:
            raise ValidationError("No se ha seleccionado una ruta de salida.")
        if self.df_bcp is None:
            raise ValidationError("Aún no se ha procesado un archivo.")
        try:
            self.df_bcp.to_excel(file_path, index=False)
            return f"✅ Archivo guardado en: {file_path}"
        except Exception as exc:
            raise ExportError(f"No se pudo exportar la plantilla BCP: {exc}") from exc
