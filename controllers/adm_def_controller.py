from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from PySide6.QtWidgets import QFileDialog

from models.process_data_def import DefProcessor

class AdministracionDefController:
    def __init__(self,view):
        self.view = view
        self.df=None
        self.file_path=""


    def getFilePath(self):
        return self.file_path
    def CargarDataFrame(self,file_path):
        try:
            self.file_path=file_path
            self.df= pd.read_excel(
                                file_path, 
                                dtype={
                                        'Cuenta Seleccionada': str,"Clasificacion Banco":str
                                        ,'N° DOCUMENTO':str,'DNI/RUC':str
                                        ,'N° CUENTA':str,'CCI':str
                                       }
                        )
        except Exception as e:
            if self.file_path=="":
                self.view.mostrarMensaje("Error al cargar el archivo", "No se ha seleccionado un archivo válido", "error")
            else: 
                self.view.mostrarMensaje("Error al cargar el archivo", str(e), "error")
            # self.mostrarMensaje("Problemas Exportando datos", str(e), "error")
            print(f"Error al cargar el archivo: {e}")
            return e
    def procesarDEF(self):
        self.df=DefProcessor(self.df).clean_data()
    def procesarBCP(self):
        self.df=DefProcessor(self.df).Plantilla()

    def analizar_dataframe_BCP(self):


        num_filas,monto_filas,num_filas_rev,monto_filas_rev,num_filas_atipicos=0,0,0,0,0
        num_filas=self.df.shape[0]
        monto_filas=self.df["IMPORTE"].sum()
        num_filas_rev=self.df[
            self.df["Doc. Verificar"].str.contains("revisar",na=False) |
            self.df["Clasificacion Banco"].str.contains("revisar",na=False)
        ].shape[0]
        monto_filas_rev=self.df[
            self.df["Doc. Verificar"].str.contains("revisar",na=False) |
            self.df["Clasificacion Banco"].str.contains("revisar",na=False)
        ]["IMPORTE"].sum()
        num_filas_atipicos=self.df[self.df["IMPORTE"]>1500].shape[0]
        
        self.procesarBCP()

        return num_filas,monto_filas,num_filas_rev,monto_filas_rev,num_filas_atipicos
    def analizar_dataframe_def(self):
        self.procesarDEF()
        num_filas,monto_filas,num_filas_rev,monto_filas_rev,num_filas_atipicos=0,0,0,0,0
        num_filas=self.df.shape[0]
        monto_filas=self.df["IMPORTE"].sum()
        num_filas_rev=self.df[
            self.df["Doc. Verificar"].str.contains("revisar",na=False) |
            self.df["Clasificacion Banco"].str.contains("revisar",na=False)
        ].shape[0]
        monto_filas_rev=self.df[
            self.df["Doc. Verificar"].str.contains("revisar",na=False) |
            self.df["Clasificacion Banco"].str.contains("revisar",na=False)
        ]["IMPORTE"].sum()

        num_filas_atipicos=self.df[self.df["IMPORTE"]>1500].shape[0]


        return num_filas,monto_filas,num_filas_rev,monto_filas_rev,num_filas_atipicos


    def ExportarDataFrame_bcp(self,file_path="export.xlsx"):
        
        if self.df is None:
            return "Aun no se ha cargado el archivo"

        try:
            fecha_actual = datetime.now().strftime("%Y%m%d%M%S")

            # Nombre del archivo con la fecha
            file_path = f"{file_path}_{fecha_actual}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
            None, "Guardar archivo", file_path, "Excel Files (*.xlsx)"
            )
            if not file_path:
                return "❌ Exportación cancelada."
                
            self.df.to_excel(file_path, index=False)

            return f"✅ Archivo guardado en: {file_path}"
        except Exception as e:
            print(f"Error al cargar el archivo: {e}")
            return e
        

    def ExportarDataFrame_def(self,file_path="export.xlsx"):
        
        if self.df is None:
            return "Aun no se ha cargado el archivo"

        try:
            fecha_actual = datetime.now().strftime("%Y%m%d%M%S")

            # Nombre del archivo con la fecha
            file_path = f"{file_path}_{fecha_actual}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
            None, "Guardar archivo", file_path, "Excel Files (*.xlsx)"
            )
            if not file_path:
                return "❌ Exportación cancelada."
                
            self.df.to_excel(file_path, index=False)

            # Aplicar colores con openpyxl
            wb = load_workbook(file_path)
            hoja = wb.active  
            fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            # Obtener índices de las columnas relevantes
            col_estado = self.df.columns.get_loc("Doc. Verificar") + 1  # +1 porque openpyxl usa índices desde 1
            col_clasificacion = self.df.columns.get_loc("Clasificacion Banco") + 1

            # Pintar los encabezados
            hoja.cell(row=1, column=col_estado).fill = fill_blue
            hoja.cell(row=1, column=col_clasificacion).fill = fill_blue
            # Autoajustar el ancho de las columnas
            for col_num, column_cells in enumerate(hoja.iter_cols(min_row=1, max_row=1), start=1):
                max_length = 0
                col_letter = get_column_letter(col_num)  

                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = max_length + 2  
                hoja.column_dimensions[col_letter].width = adjusted_width


            # Guardar cambios
            wb.save(file_path)
            wb.close()


            return f"✅ Archivo guardado en: {file_path}"
        






        except Exception as e:
            print(f"Error al cargar el archivo: {e}")
            return e

    
    
    
    
    def validate_columns_DEF(self):
        REQUIREMENT_COLUMNS=['N° NOTA CRÉDITO',
                        'FECHA NCR',
                        'TIPO DOC.',
                        'N° DOCUMENTO',
                        'DNI/RUC',
                        'NOMBRE CLIENTE/RAZÓN SOCIAL',
                        'IMPORTE',
                        'FECHA REGISTRO DATOS',
                        'CUENTA TITULAR',
                        'N° CUENTA',
                        'CCI',
                        'BANCO',
                        'ESTADO',
                        'CORREO '
                        ]
        
        missing_columns = [col for col in REQUIREMENT_COLUMNS if col not in self.df.columns]
        
        if missing_columns:
            self.df=None
            raise ValueError(f"❌ Faltan las siguientes columnas: {', '.join(missing_columns)}")
        

        
    def validate_columns_BCP(self):
        REQUIREMENT_COLUMNS=['Clasificacion Banco',
                        'Cuenta Seleccionada',
                        'Clasificacion Doc',
                        'DNI/RUC',
                        'NOMBRE CLIENTE/RAZÓN SOCIAL',
                        'IMPORTE',
                        'DEF FORMATEADO',
                        ]
        
        missing_columns = [col for col in REQUIREMENT_COLUMNS if col not in self.df.columns]
        
        if missing_columns:
            self.df=None
            raise ValueError(f"❌ Faltan las siguientes columnas: {', '.join(missing_columns)}")
        

